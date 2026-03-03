"""
Manual Excel database collector.

Reads Base de datos maual/Metabolitos.xlsx and imports individual
metabolite entries into the DB.  Category/panel entries are skipped.

The Excel has four 'Biomarcador' columns:
  G  (col 7)  — non-GC biomarkers (others methods)
  O  (col 15) — GC biomarkers (list)
  X  (col 24) — GC biomarkers from Nerea's study (Class – Compound format)
  AE (col 31) — GC biomarkers (list)

All entries are tagged:
  - schizophrenia=True, mental_health=True
  - fecal_hint=True (fecal GC-MS context)
  - method_hint="GC" for GC columns (O, X, AE)
  - source_type="Manual_MH"

Returns (n_links, health_dict).
"""

from __future__ import annotations
import logging
import re
from pathlib import Path
from typing import Any

from src.db import (
    get_conn, upsert_metabolite, upsert_source,
    add_synonym, link_metabolite_source, migrate_db,
)
from src.normalize import normalize, make_key

logger = logging.getLogger(__name__)

# Biomarker column indices (1-based → we use 0-based in tuple access)
_COLS = {
    "G":  (6,  None),   # col G, no method hint (non-GC methods)
    "O":  (14, "GC"),   # col O, GC
    "X":  (23, "GC"),   # col X, GC Nerea
    "AE": (30, "GC"),   # col AE, GC
}

# ── Skip patterns ──────────────────────────────────────────────────────────────
_SKIP_SUBSTRINGS = ["etc.", " + ", "↓", "↑", "...", "panel", "perfil "]
_SKIP_PREFIXES_LC = [
    "scfa", "aminoácidos,", "alcanos/alquenos", "aldehídos ",
    "amplio rango", "vocs discrimin", "ácidos grasos de cadena",
    "cetonas (energía", "fosfatidiletanolamina", "fosfatidilglicerol",
    "fosfatidilserina", "ácido fosfatídico", "apolipoproteína",
    "lisofosfatidilcolinas", "diacilglicérido", "triacilglicérido",
    "monoacilglicérido", "triacilglicerol", "plasmalógenos",
    "ceramidas", "éster de colesterol",
    "aminoácidos, aminas",
]


def _should_skip(name: str) -> bool:
    nl = name.strip().lower()
    for s in _SKIP_SUBSTRINGS:
        if s in name:
            return True
    for p in _SKIP_PREFIXES_LC:
        if nl.startswith(p):
            return True
    return False


# Regex to detect English chemical names worth using as canonical form
_ENGLISH_CHEM_RE = re.compile(
    r"\b(?:acid|alcohol|acetate|propionate|butyrate|valerate|caproate|caproic|"
    r"isobutyric|isovaleric|isocaproic|isohexanoic|indole|skatole|phenol|"
    r"sulfide|trisulfide|disulfide|methanethiol|dimethyl|pentanone|hexanone|"
    r"pentanol|hexanol|butanal|pentanal|nonanal|benzaldehyde|methional)\b",
    re.IGNORECASE,
)


def _parse_entry(raw: str) -> tuple[str, list[str]]:
    """
    Parse a single Excel cell and return (canonical_name, [synonyms]).
    Returns ("", []) if the entry should be skipped.
    """
    entry = (raw or "").strip().rstrip(".")
    if not entry or _should_skip(entry):
        return "", []

    synonyms: list[str] = []

    # Pattern 1: "Class \u2013 Compound (opt_ABBREV)"
    for sep in (" \u2013 ", " \u2014 "):
        if sep in entry:
            compound = entry.split(sep, 1)[1].strip()
            # Strip trailing short ALL-CAPS abbreviation like "(DMTS)"
            m = re.match(r"^(.+?)\s+\(([A-Z]{2,6})\)\s*$", compound)
            if m:
                synonyms.append(m.group(2))
                compound = m.group(1).strip()
            if not compound or _should_skip(compound) or "..." in compound:
                return "", []
            return compound, synonyms

    # Pattern 2: "Name (alias/context)" — may have multiple parens
    parens = re.findall(r"\(([^)]+)\)", entry)
    base   = re.sub(r"\s*\([^)]*\)", "", entry).strip().rstrip(".")

    if not base:
        return "", []

    for p in parens:
        p = p.strip()
        # English chemical name in parens → prefer it as canonical
        if _ENGLISH_CHEM_RE.search(p) and len(p) < 60:
            synonyms.append(base)
            base = p
            break
        # Short ALL-CAPS abbreviation → make synonym
        if re.match(r"^[A-Z]{2,6}$", p):
            synonyms.append(p)

    if not base or _should_skip(base):
        return "", []

    return base, synonyms


def collect(cfg: dict, db_path: str) -> tuple[int, dict]:
    """
    Import manual Excel biomarkers into the DB.
    Returns (n_links, health_dict).
    """
    migrate_db(db_path)

    excel_path = Path(cfg.get("manual_excel_path",
                              "Base de datos maual/Metabolitos.xlsx"))
    if not excel_path.exists():
        logger.warning("manual_excel: file not found at %s", excel_path)
        return 0, {"error": "file not found", "added": 0}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.error("manual_excel: could not open %s: %s", excel_path, exc)
        return 0, {"error": str(exc), "added": 0}

    health: dict[str, Any] = {
        "rows_scanned":   0,
        "entries_parsed": 0,
        "entries_skipped": 0,
        "links_added":    0,
        "last_error":     "",
    }

    # Tags for all manual entries (MH biomarkers — schizophrenia focus)
    _BASE_TAGS = {
        "mental_health":  True,
        "schizophrenia":  True,
        "fecal_hint":     True,
        "mh_biomarker":   True,
        "condition_hits": ["schizophrenia"],
    }

    n_links = 0

    with get_conn(db_path) as conn:
        for data_row in ws.iter_rows(min_row=22, max_row=ws.max_row, values_only=True):
            health["rows_scanned"] += 1

            for col_label, (col_idx, method_hint) in _COLS.items():
                raw = data_row[col_idx]
                if raw is None:
                    continue
                raw_str = str(raw).strip()
                if not raw_str:
                    continue

                canon_name, synonyms = _parse_entry(raw_str)
                if not canon_name:
                    health["entries_skipped"] += 1
                    continue

                health["entries_parsed"] += 1

                # Upsert source for this column group
                src_id = upsert_source(
                    conn,
                    source_type="Manual_MH",
                    source_ref=f"Excel_{col_label}",
                    title="Manual database - fecal GC-MS mental health biomarkers",
                    matrix_hint="fecal",
                    method_hint=method_hint,
                )

                nkey = make_key(canon_name)
                mid  = upsert_metabolite(
                    conn, normalize(canon_name), nkey, tags=_BASE_TAGS,
                )
                add_synonym(conn, mid, canon_name, nkey)
                for syn in synonyms:
                    add_synonym(conn, mid, syn, make_key(syn))

                link_metabolite_source(
                    conn, mid, src_id,
                    evidence_tag="manual_curation",
                )
                n_links += 1

    health["links_added"] = n_links
    logger.info(
        "manual_excel: scanned=%d parsed=%d skipped=%d links=%d",
        health["rows_scanned"], health["entries_parsed"],
        health["entries_skipped"], n_links,
    )
    return n_links, health
